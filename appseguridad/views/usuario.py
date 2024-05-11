from django.db.models import Q 
from django.contrib import messages
from django.shortcuts import render, redirect,  get_object_or_404
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from appevaluacion.forms import *
from appevaluacion.models import *
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from django.contrib.auth.hashers import is_password_usable
from django.conf import settings
from appseguridad.forms import CrearUserForm,EditarUserForm, PerfilrUserForm, CorreoForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from openpyxl import Workbook
from django.http.response import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import openpyxl
from openpyxl import Workbook
from django.core.mail import EmailMessage

#---------------------------------------------Listar usuarios---------------------------------------------
@login_required(login_url='login')
def listar_usuarios(request):
    return render(request, 'usuario/listar.html')

@login_required(login_url='login')
def listar_usuarios_json(request):    
    filtro_status = request.GET.get('estado')
    if filtro_status == "activo":
        usuarios = list(User.objects.filter(is_active=True).values())
    elif filtro_status == "inactivo":
        usuarios = list(User.objects.filter(is_active=False).values())
    else:
        usuarios = list(User.objects.values())
    data = {'usuarios':usuarios}
    return JsonResponse(data)

# Actualiza estado del Usuario
@csrf_exempt
def update_user_status(request):
    user_id = request.POST.get('user_id')
    is_active = request.POST.get('is_active') == 'true'  
    try:
        user = User.objects.get(id=user_id)
        user.is_active = is_active
        user.save()
        return JsonResponse({'status': 'success'})
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
    
#---------------------------------------------------------------------------------------------------------

#------------------------------------------Crer USuario --------------------------------------------------
@login_required(login_url='login')
def creacion_usuario(request):
    if request.method == 'POST':
        form = CrearUserForm(request.POST)
        if form.is_valid():
            form_instance = form.save(commit=False)
            
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            evaluador = form.cleaned_data['first_name']
            email = form.cleaned_data['email']
            form_instance.save() 
            id = form_instance.id
            save_password(id, username, password, evaluador, email)
            
            #form.save()
            messages.success(request, f'Usuario {username} creado exitosamente.')
            return redirect('listar_usuarios')
        else:
            messages.error(request, "El formulario contiene errores. Por favor, corrija")
    else:
        form = CrearUserForm()
    return render(request, 'usuario/agregar.html', {'form': form})

#Guardar en un archivo excel las credenciales
def save_password(id, username, password, evaluador, email):
    # Ruta del archivo Excel
    excel_file = 'C:/Users/mijha/Desktop/modulo curricular/proypostgrado/appseguridad/templates/usuario/credenciales_usuarios - copia.xlsx'
    
    # Si el archivo no existe, crearlo y añadir cabeceras
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Id'
        sheet['B1'] = 'Username'
        sheet['C1'] = 'Password'
        sheet['D1'] = 'Evaluador'
        sheet['E1'] = 'Email'
        workbook.save(filename=excel_file)
    
    # Abrir el archivo existente y agregar una nueva entrada
    workbook = openpyxl.load_workbook(filename=excel_file)
    sheet = workbook.active
    max_row = sheet.max_row + 1
    sheet.cell(row=max_row, column=1).value = id
    sheet.cell(row=max_row, column=2).value = username
    sheet.cell(row=max_row, column=3).value = password
    sheet.cell(row=max_row, column=4).value = evaluador
    sheet.cell(row=max_row, column=5).value = email
    workbook.save(filename=excel_file)

#Descargar Credenciales en excel
def download_excel(request):   
    # Ruta al archivo Excel en el directorio de medios
    excel_path = 'C:/Users/mijha/Desktop/modulo curricular/proypostgrado/appseguridad/templates/usuario/credenciales_usuarios - copia.xlsx'
    # Respuesta de descarga del archivo Excel
    with open(excel_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=credenciales_usuarios.xlsx'
        return response

#---------------------------------------------------------------------------------------------------------

#------------------------------------------Editar USuario ------------------------------------------------
@login_required(login_url='login')
def actualizar_usuario(request, id):
    usuario = get_object_or_404(User, pk=id)
    if request.method == 'POST':
        form = EditarUserForm(request.POST, instance=usuario)
        if form.is_valid():
           form.save()
           username = form.cleaned_data.get('username')
           messages.success(request, f'Usuario {username} actualizado exitosamente.')
           return redirect('listar_usuarios')
        else:
           messages.error(request, "El formulario contiene errores. Por favor, corrija")
    else:
        form = EditarUserForm(instance=usuario)
    return render(request, 'usuario/editar.html', {'form': form})

#---------------------------------------------------------------------------------------------------------

#------------------------------------------Eliminar USuario ------------------------------------------------
@login_required
def eliminar_usuario(request, id):
    usuario = get_object_or_404(User, pk=id)   
    if request.user.is_superuser:  
        usuario.delete()
        messages.success(request, 'Usuario eliminado exitosamente.')
    else:
        messages.error(request, 'No tienes permiso para eliminar usuarios.')
    return redirect('listar_usuarios') 

#---------------------------------------------------------------------------------------------------------

#------------------------------------------Editar Perfil USuario ------------------------------------------
@login_required
def editar_perfil(request):
    if request.method == 'POST':
        form = PerfilrUserForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado.')
            return redirect('home')
    else:
        form = PerfilrUserForm(instance=request.user)

    return render(request, 'usuario/perfil.html', {'form': form})

#---------------------------------------------------------------------------------------------------------

#------------------------------------------ Envio de correos -----------------------------------------
@login_required(login_url='login')
def copiar_usuarios_activos():
    # Ruta del archivo Excel original y el archivo de destino
    excel_file_origen = 'C:/Users/mijha/Desktop/modulo curricular/proypostgrado/appseguridad/templates/usuario/credenciales_usuarios - copia.xlsx'
    excel_file_destino = 'C:/Users/mijha/Desktop/modulo curricular/proypostgrado/appseguridad/templates/usuario/usuarios_activos.xlsx'
    
    # Si el archivo de destino no existe, crearlo y añadir cabeceras
    if not os.path.exists(excel_file_destino):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Id'
        sheet['B1'] = 'Username'
        sheet['C1'] = 'Password'
        sheet['D1'] = 'Evaluador'
        sheet['E1'] = 'Email'
        workbook.save(filename=excel_file_destino)
    
    # Abrir el archivo Excel original y el archivo de destino
    workbook_origen = openpyxl.load_workbook(filename=excel_file_origen)
    sheet_origen = workbook_origen.active
    
    workbook_destino = openpyxl.load_workbook(filename=excel_file_destino)
    sheet_destino = workbook_destino.active

    usuarios_activos = User.objects.filter(is_active=True)
    ids_activos = [usuario.id for usuario in usuarios_activos]
    
    # Depurar: Imprimir los IDs activos y compararlos con los IDs en el archivo Excel
    print("IDs activos en la base de datos:", ids_activos)
    
       # Copiar los IDs de los usuarios activos al archivo de destino
    for index, user_id in enumerate(ids_activos, start=1):
        sheet_destino.cell(row=index, column=1).value = user_id
    
    # Guardar los cambios en el archivo de destino
    workbook_destino.save(filename=excel_file_destino)



#Descargar Credenciales en excel
def download_excel_activos(request):   
    # Ruta al archivo Excel en el directorio de medios
    excel_path = 'C:/Users/mijha/Desktop/modulo curricular/proypostgrado/appseguridad/templates/usuario/usuarios_activos.xlsx'
    # Respuesta de descarga del archivo Excel
    with open(excel_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=usuarios_activos.xlsx'
        return response

@login_required(login_url='login')
def enviar_correo(request):
    if request.method == 'POST':
        form = CorreoForm(data = request.POST)
        if form.is_valid():
            nombre = form.cleaned_data.get('nombre')
            email = form.cleaned_data.get('email')
            contenido = form.cleaned_data.get('contenido')

            email = EmailMessage("Mensaje del Servidor",
                    "El usuario con nombre {} con la direccion {} escribe lo siguiente: \n\n {}".format(nombre,email,contenido),'', ['mijharv@gmail.com', 'mjrojasv@unitru.edu.pe'], reply_to=[email])
            
            try:
                email.send()
                print('Se envio correctamente')
            except Exception as e:
                print(f'No se envio correctamente : {e}')

            return redirect('listar_usuarios')      
    else:
        form = CorreoForm()
    return render(request, 'usuario/envio_correo.html', {'form': form})



#---------------------------------------------------------------------------------------------------------
